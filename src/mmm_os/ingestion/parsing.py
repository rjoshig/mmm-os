"""Low-level file readers for CSV and multi-tab XLSX (P1-2).

Reads are streamed/chunked (P1-6): structure detection uses a bounded row preview,
and profiling (01.3) iterates a single sheet lazily. Cells are normalised to
``str | None`` so CSV and XLSX are handled uniformly downstream.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import PurePosixPath
from typing import IO

from openpyxl import load_workbook

Cell = str | None
Row = list[Cell]


class ParseError(RuntimeError):
    """Raised when a file cannot be parsed."""


class UnsupportedFileTypeError(ParseError):
    """Raised for a file extension we do not support yet."""


@dataclass
class RawSheet:
    """A bounded preview of one sheet's rows.

    Attributes:
        name: The sheet name.
        index: The zero-based sheet index within the workbook.
        rows: A bounded list of normalised rows (each a list of ``str | None``).
        truncated: Whether more rows exist beyond the preview.
    """

    name: str
    index: int
    rows: list[Row] = field(default_factory=list)
    truncated: bool = False

    def is_empty(self) -> bool:
        """Return whether the previewed sheet has no non-empty cell."""
        return all(cell is None for row in self.rows for cell in row)


def _normalize(value: object) -> Cell:
    """Normalise a raw cell value to a trimmed string or ``None``."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


@dataclass(frozen=True)
class FixedWidthField:
    """One column in a fixed-width layout.

    Attributes:
        name: The column name (used as the synthesized header).
        start: Zero-based, inclusive start offset within each line.
        width: Number of characters the field spans.
    """

    name: str
    start: int
    width: int


@dataclass(frozen=True)
class ParseOptions:
    """How to parse a text feed when the extension alone is not enough (Slice 7.4).

    ``fmt="auto"`` selects by extension (``.xlsx`` → xlsx, else delimited with a
    sniffed delimiter). ``delimited`` reads a character-separated file (``delimiter``
    given, else sniffed). ``fixed_width`` slices each line by ``fixed_fields``,
    synthesizing a header row from the field names.
    """

    fmt: str = "auto"
    delimiter: str | None = None
    has_header: bool = True
    fixed_fields: tuple[FixedWidthField, ...] = ()


# Clearly-tabular text extensions the generic (no-template) path accepts. Ambiguous
# extensions (.txt/.dat) are only parsed when a feed template supplies ParseOptions.
_DELIMITED_SUFFIXES = {".csv", ".tsv", ".psv"}


def file_kind(filename: str) -> str:
    """Return the parser kind for a filename (``"csv"`` or ``"xlsx"``).

    Args:
        filename: The file's name.

    Returns:
        ``"csv"`` for any delimited-text extension, ``"xlsx"`` for workbooks.

    Raises:
        UnsupportedFileTypeError: For any other extension.
    """
    suffix = PurePosixPath(filename).suffix.lower()
    if suffix in _DELIMITED_SUFFIXES:
        return "csv"
    if suffix == ".xlsx":
        return "xlsx"
    raise UnsupportedFileTypeError(f"unsupported file type: {suffix or filename!r}")


def _kind_for(filename: str, options: ParseOptions | None) -> str:
    """Return ``"csv"`` (text) or ``"xlsx"`` for a read.

    With a feed template (``options``), the declared format wins regardless of the
    filename extension — so ``.txt``/``.dat`` fixed-width and delimited feeds parse.
    Without one, the extension must be a known tabular/workbook type.
    """
    if options is not None:
        return "xlsx" if options.fmt == "xlsx" else "csv"
    return file_kind(filename)


def _resolve_delimiter(filename: str, options: ParseOptions | None) -> str | None:
    """Return an explicit delimiter for a delimited read, or ``None`` to sniff.

    An explicit ``options.delimiter`` wins; otherwise a ``.tsv``/``.psv`` extension
    implies its delimiter; a plain ``.csv`` implies a comma; everything else sniffs.
    """
    if options is not None and options.delimiter:
        return options.delimiter
    suffix = PurePosixPath(filename).suffix.lower()
    if suffix == ".tsv":
        return "\t"
    if suffix == ".psv":
        return "|"
    if suffix == ".csv":
        return ","
    return None


def _sniff_delimiter(sample: str) -> str:
    """Best-effort detect a delimiter from a text sample (comma fallback)."""
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        return ","


def _delimited_rows(stream: IO[bytes], delimiter: str | None) -> Iterator[Row]:
    text = io.TextIOWrapper(stream, encoding="utf-8-sig", newline="")
    if delimiter is None:
        sample = text.read(8192)
        text.seek(0)
        delimiter = _sniff_delimiter(sample)
    for raw in csv.reader(text, delimiter=delimiter):
        yield [_normalize(cell) for cell in raw]


def _fixed_width_rows(
    stream: IO[bytes], fields: tuple[FixedWidthField, ...], has_header: bool
) -> Iterator[Row]:
    """Yield rows from a fixed-width file, synthesizing a header from field names."""
    if not fields:
        raise ParseError("fixed_width parsing requires at least one field")
    yield [f.name for f in fields]
    text = io.TextIOWrapper(stream, encoding="utf-8-sig", newline="")
    for i, line in enumerate(text):
        if has_header and i == 0:
            continue  # the source's own header line is discarded
        stripped = line.rstrip("\r\n")
        if not stripped.strip():
            continue
        yield [_normalize(stripped[f.start : f.start + f.width]) for f in fields]


def _text_rows(stream: IO[bytes], filename: str, options: ParseOptions | None) -> Iterator[Row]:
    """Dispatch a text (non-xlsx) read to the delimited or fixed-width reader."""
    if options is not None and options.fmt == "fixed_width":
        yield from _fixed_width_rows(stream, options.fixed_fields, options.has_header)
        return
    yield from _delimited_rows(stream, _resolve_delimiter(filename, options))


def _csv_rows(stream: IO[bytes]) -> Iterator[Row]:
    yield from _delimited_rows(stream, ",")


def read_preview(
    stream: IO[bytes],
    filename: str,
    max_rows: int,
    options: ParseOptions | None = None,
) -> list[RawSheet]:
    """Read a bounded preview of every sheet for structure detection.

    Args:
        stream: A readable binary stream of the file.
        filename: The file's name (selects the parser).
        max_rows: Maximum rows to read per sheet.
        options: Optional parse options (delimiter / fixed-width) from a feed
            template; when ``None`` the format is inferred from the extension.

    Returns:
        One ``RawSheet`` per sheet (all sheets for XLSX; a single sheet for text).
    """
    kind = _kind_for(filename, options)
    if kind == "csv":
        rows: list[Row] = []
        truncated = False
        for i, row in enumerate(_text_rows(stream, filename, options)):
            if i >= max_rows:
                truncated = True
                break
            rows.append(row)
        return [
            RawSheet(
                name=PurePosixPath(filename).stem or "Sheet1",
                index=0,
                rows=rows,
                truncated=truncated,
            )
        ]

    workbook = load_workbook(stream, read_only=True, data_only=True)
    try:
        sheets: list[RawSheet] = []
        for index, worksheet in enumerate(workbook.worksheets):
            rows = []
            truncated = False
            for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                if i >= max_rows:
                    truncated = True
                    break
                rows.append([_normalize(cell) for cell in row])
            sheets.append(
                RawSheet(name=worksheet.title, index=index, rows=rows, truncated=truncated)
            )
        return sheets
    finally:
        workbook.close()


def iter_sheet_rows(
    stream: IO[bytes],
    filename: str,
    sheet_index: int,
    options: ParseOptions | None = None,
) -> Iterator[Row]:
    """Iterate all rows of a single sheet lazily (for profiling).

    Args:
        stream: A readable binary stream of the file.
        filename: The file's name (selects the parser).
        sheet_index: The zero-based sheet index to iterate.
        options: Optional parse options (delimiter / fixed-width) from a feed
            template; when ``None`` the format is inferred from the extension.

    Yields:
        Normalised rows of the requested sheet.
    """
    kind = _kind_for(filename, options)
    if kind == "csv":
        yield from _text_rows(stream, filename, options)
        return

    workbook = load_workbook(stream, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[sheet_index]
        for row in worksheet.iter_rows(values_only=True):
            yield [_normalize(cell) for cell in row]
    finally:
        workbook.close()
